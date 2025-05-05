import subprocess
import os

import difflib


import tkinter as tk
from tkinter import ttk

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    from openpyxl.styles import Font
except:
    print("\nVerifique que tiene instaladas las librerías openpyxl y pandas.")

    
import socket
import hashlib
import platform
from datetime import datetime

programa = 'python3' if os.name == "posix" else 'python'

# archivo de salida
nombre = "Resultados_Estudiante.xlsx"
time = 5 # segundos máximos de ejecución de cada test

dict_errores = {
    'Semántica': "    El código presentado está correctamente escrito en \
Python.\n    Pero tiene problemas con el planteamiento del ejercicio y el \
cómo resolverlo.\n    (Esto puede deberse a:: \n        El formato requerido \
en el ejercicio no se ha respetado completamente.\n        El código está \
incompleto.\n        La solución no corresponde al ejercicio \
solicitado.\n    )",
    'IndexError': "    Acceso a posiciones inexistentes en una colección \
(Listas, Strings).\n    Ejemplo:\nacceder al quinto elemento de una lista que \
solo tiene 2 elementos.",
    'TypeError': "    Uso de una operación con un tipo de dato incompatible. \
   Ejemplo (sumar un número y un texto):\n    5 + 'a' ",
    'NameError': "    Uso de elementos que no se encuentran definidos \
previamente en el código.\n    Ejemplo:\n    Usar la variable hola antes de \
asignarle un valor.",
    'ValueError': "    Se ha utilizado un valor inesperado para la \
operación.\n    Ejemplo:\n    Transformar una palabra en un número.\n    \
int('ABC') ",
    'AttributeError': "    Confusión de las propiedades de los tipos de datos \
o uso de una propiedad inexistente para un tipo de dato.\n    Ejemplo:\n    \
Uso de elementos propios de los string en una lista.\n    lista.split(';')",
    'SyntaxError': "    El código no está utilizando correctamente la \
escritura utilizada en Python\n    Ejemplo:\n    (25 * 3) +  8)\n    \
(falta/sobra un paréntesis)",
    'EOFError': "    Incompatibilidad entre la cantidad de datos que necesita \
el programa y los entregados.\n    Ejemplo:\n    Pedir más/menos entradas de \
las que el programa luego utiliza.",
    'IndentationError': "    El programa no respeta las tabulaciones \
(sangría) que necesita la sintaxis de Python.\n    Ejemplo:\n    No \
escribir más a la derecha luego de un if.",
    'Ciclo infinito': "    Existe algún ciclo que tarda más de lo esperado en \
terminar. Probablemente debido a una tautología.\n    Ejemplo:\n    No \
actualizar la condición de un ciclo while para que en algún punto se vuelva \
falsa.",
    'UnboundLocalError': "    Una variable local fue utilizada antes de ser \
definida.",
    'FileNotFoundError': "    Un archivo solicitado no ha sido encontrado.",
    'Success': "    Test superado con éxito.",
    'UnicodeEncodeError':"Error de codificación de algunos caracteres. Si no sabe a que se debe, consulte con alguno de los docentes."
}


def escribir_multi(name, df_list, hoja_list):
    # Escribir el DataFrame en un nuevo archivo Excel
    try:

        
        with pd.ExcelWriter(nombre, engine='openpyxl') as writer:
            for i in range(len(df_list)):
                df = df_list[i]
                hoja = hoja_list[i]
                df.to_excel(writer, index=False, sheet_name=hoja)
                # writer.save()

        # Cargar el libro de trabajo de Excel
        wb = load_workbook(nombre)

        # Obtener la hoja de cálculo
        hoja = wb["Test"]

        # Configurar la opción wrap_text en True para todas las celdas
        for fila in hoja.iter_rows():
            for celda in fila:
                celda.alignment = Alignment(wrapText=True)

        # Guardar los cambios
        wb.save(name)
    except:
        
        print("\nEl archivo Resultados_Estudiante.xlsx puede estar abierto.")
        print("Por favor, ciérralo antes de continuar.")


# Entradas y Salidas
try:
    with open("IO_Esperado.txt", encoding="utf-8") as a:
        lista = a.readlines()
except UnicodeDecodeError:
    with open("IO_Esperado.txt", encoding="cp1252") as a:
        lista = a.readlines()

testE = []
testS = []
io = "###ENTRADA###"
data = []
for linea in lista:
    if "###ENTRADA###" == linea.strip():
        testS.append(data)
        data = []
        io = "###ENTRADA###"
    elif "###SALIDA###" == linea.strip():
        testE.append(data)
        data = []
        io = "###SALIDA###"
    else:
        data.append(linea.strip())
testS.append(data)
testS.pop(0)
data = []

# archivos a analizar
alumnos = os.listdir('./estudiantes')
cantidad = str(len(alumnos))

df_list = []

list_test = []
for i in range(1, len(testE)+1):
    list_test.append("TEST "+"0"*(2-len(str(i)))+str(i))
    list_test.append("Entrada "+str(i))
    list_test.append("Salida Estudiante "+str(i))
    list_test.append("Salida Esperada "+str(i))
    list_test.append("Situación Test "+str(i))
list_test.append("Correctas")

# Aquí empieza
actual = 1
correctos = []
fallidos = []
for alumno in alumnos:

    lista_linea = []

    if alumno[-3]+alumno[-2]+alumno[-1] == ".py":
        formato = True
    else:
        formato = False

    if not formato:
        print("Skip de archivo por formato incorrecto (debe ser un .py)")

        for e in range(len(testE)):
            lista_linea.append(0)
            lista_linea.append("")
            lista_linea.append("")
            lista_linea.append("")
        lista_linea.append(0)

    else:
        code = ''
        try:
            with open('./estudiantes/'+alumno, 'r', encoding="cp1252") as file:
                code = file.read()
        except UnicodeDecodeError:
            with open('./estudiantes/'+alumno, 'r', encoding="utf8") as file:
                code = file.read()

        name = socket.gethostname()
        add = socket.gethostbyname(name)
        code_out = '######### '+alumno
        code_out += '\n######### ' + name
        code_out += '\n######### ' + add
        code_out += '\n######### ' + str(platform.system())
        code_out += '\n######### ' + str(platform.version())
        code_out += '\n######### ' + str(platform.architecture())
        code_out += '\n######### ' + str(platform.node())
        code_out += '\n######### ' + str(platform.processor())
        code_out += '\n######### '+ str(datetime.now())
        code_out += "\n" + code
        code_out += "\n#########"
        code = code_out.encode('utf-8')
        hh = hashlib.sha256(code).hexdigest()
        code_out += "\n"+hh

        with open("log.out", 'a') as file:
             file.write("\n\n"+code_out)

        correctas = 0
        i = 0
        info_test = []
        while i < len(testE):
            try:
                proceso = subprocess.Popen(
                    [programa, './estudiantes/'+alumno],
                    stdin=subprocess.PIPE,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    text=True,
                    creationflags = subprocess.CREATE_NO_WINDOW if os.name != "posix" else 0
                    )
                entrada = "\n".join(testE[i])
                salida, error = proceso.communicate(entrada, timeout=time)

            except subprocess.TimeoutExpired:
                proceso.terminate()
                salida, error = "", "Tiempo de ejecución excedido ("+str(time)+"s)"
            salida_esperada = "\n".join(testS[i])

            #print(error)
            if error:
                #print(f"    Se produjo un error: {error}")
                lista_linea.append("0")
                lista_linea.append(entrada)
                lista_linea.append(error)
                lista_linea.append(salida_esperada)
                fallidos.append(str(i+1))
                if "Tiempo de ejecución excedido" in error or "stack overflow" in error:
                    info_test.append("Ciclo infinito")
                    lista_linea.append("Ciclo infinito: "+dict_errores["Ciclo infinito"])
                else:
                    info_test.append(error.split("\n")[-2].split(":")[0])
                    type_error = error.split("\n")[-2].split(":")[0]
                    try:
                        situacion_error = type_error+": "+dict_errores[type_error]
                    except:
                        situacion_error = type_error+": "+type_error
                    lista_linea.append(situacion_error)
                    

            else:
                resultado = salida.strip()

                if resultado == salida_esperada:
                    lista_linea.append("1")
                    correctas +=1
                    info_test.append("Success")
                    correctos.append(str(i+1))
                    situacion = "Correcta"

                else:
                    lista_linea.append("0")
                    info_test.append("Semántica")
                    fallidos.append(str(i+1))

                    diferencias = difflib.unified_diff(
                        salida_esperada.splitlines(), 
                        resultado.splitlines(), 
                        fromfile='salida_esperada', 
                        tofile='resultado', 
                        lineterm=''
                    )
                    situacion = []
                    num_ln = 0
                    ln_comp = 0
                    list_resultado = resultado
                    list_salida_esperada = salida_esperada
                    for ln in diferencias:
                        if num_ln > 2:
                            situacion.append(ln) 
                        num_ln += 1
                        
                    situacion.append('''
#############
- Indica que una línea está en la salida esperada, pero no en su resultado.
+ Indica que una línea está en su resultado, pero no en salida esperada.''')
                    
                    ln_comp = 0
                    list_resultado = resultado.split("\n")
                    list_salida_esperada = salida_esperada.split("\n")
                    dice = ""
                    decir = ""
                    ln_resultado = []
                    ln_esperada = []
                    diff_flag = False
                    for ln in range(min(len(list_resultado),len(list_salida_esperada))):
                        if list_resultado[ln] != list_salida_esperada[ln]:
                            ln_resultado =  list_resultado[ln].split(" ")
                            ln_esperada =  list_salida_esperada[ln].split(" ")
                            for pal in range( min(len(ln_resultado),len(ln_esperada)) ):
                                if ln_resultado[pal] != ln_esperada[pal]:
                                    dice = ln_resultado[pal]
                                    decir = ln_esperada[pal]
                                    diff_flag = True
                                    break
                            if diff_flag:
                                break
                                    
                    situacion.append( "#############\n")
                    if len(ln_resultado) < len(ln_esperada):
                        situacion.append("Existe un espacio extra.")
                    elif len(ln_resultado) > len(ln_esperada):
                        situacion.append("El formato de espaciado no corresponde.")
                    else:
                        situacion.append( "Dice: "+ dice )            
                        situacion.append( "Debería decir: "+ decir )            
                    situacion = "\n".join(situacion)

                lista_linea.append(entrada)
                lista_linea.append(resultado)
                lista_linea.append(salida_esperada)
                lista_linea.append(situacion)

            i+=1

        lista_linea.append(str(correctas))
        

    try:
        print(str(actual)+"/"+cantidad +" "+alumno+": "+str(correctas)+"/"+str(len(testS)))
        print()
        for e in range(len(info_test)):
        
            if info_test[e] == "Success":
                print("Test "+"0"*(2-len(str(e+1)))+str(e+1)+": "+dict_errores[info_test[e]])
            else :
                print("\nTest "+"0"*(2-len(str(e+1)))+str(e+1)+": Error de "+info_test[e])
                print(dict_errores[info_test[e]])

        print()
        df_list.append(lista_linea)
        actual += 1
    
    except:
        print("El archivo no existe o tiene formato incorrecto.")

try:
    df = pd.DataFrame(df_list , columns = list_test)
    data_frames_lista = [df]
    hojas_lista = ['Test']
        
    escribir_multi(nombre,data_frames_lista,hojas_lista)
except:
    print("...")
    
print("\n\n>>>>>>   COMPLETADO...   <<<<<<<\n\n")
print("    Test correctos: " + " ".join(correctos))
print("    Test fallidos: " + " ".join(fallidos))


################################################################################
################################################################################
################################################################################
################################################################################
################################################################################
####################
####################
####################


headers = ['Test', 'Entrada', 'Salida Estudiante', 'Salida Esperada', 'Situación']


# Crear ventana principal
root = tk.Tk()
root.title("Visor de Tests")
root.geometry("1400x600")

# Función para cargar los datos del test seleccionado
def cargar_test(event):
    seleccion = combobox.get()  
    indice = int(seleccion.split()[1]) - 1  
    inicio = indice * 5  # indice inicial
    test_data = df_list[0][inicio:inicio + 5]  # cada 5

    # limpiar 
    for widget in frame_contenedor.winfo_children():
        widget.destroy()

    # mostrar Test y valor (Correcto o Fallido)
    test_result = "Correcto" if test_data[0] == "1" else "Fallido"
    tk.Label(root, text=str(seleccion)+":", font=("Arial", 12, "bold")).place(x=10, y=10)  
    tk.Label(root, text=test_result, font=("Arial", 12), width=8).place(x=90, y=10)  

    # Mostrar CORRECTOS y FALLIDOS en la esquina superior derecha
    
    tk.Label(root, text="TEST CORRECTOS:", font=("Arial", 12, "bold")).place(x=900, y=10)
    tk.Label(root, text=" ".join(correctos), font=("Arial", 12), width=15).place(x=1060, y=10)

    tk.Label(root, text="TEST FALLIDOS:", font=("Arial", 12, "bold")).place(x=900, y=40)
    tk.Label(root, text=" ".join(fallidos), font=("Arial", 12), width=15).place(x=1060, y=40)

    row_index = 1  

    # Mostrar Entrada y su valor
    tk.Label(frame_contenedor, text="Entrada", font=("Arial", 12, "bold")).grid(row=row_index, column=0, padx=5, pady=5, sticky="w")
    text_widget_entrada = tk.Text(frame_contenedor, wrap="word", width=80) 
    text_widget_entrada.insert("1.0", test_data[1])
    text_widget_entrada.configure(state="disabled")

    # ajustar la altura 
    height = len(test_data[1].split("\n"))
    text_widget_entrada.configure(height=height)
    text_widget_entrada.grid(row=row_index, column=1, columnspan=5, padx=5, pady=5, sticky="ew")

    row_index += 1  

    # Salida Estudiante y Salida Esperada 
    tk.Label(frame_contenedor, text="Salida Estudiante", font=("Arial", 12, "bold")).grid(row=row_index, column=0, padx=5, pady=5, sticky="w")
    tk.Label(frame_contenedor, text="Salida Esperada", font=("Arial", 12, "bold")).grid(row=row_index, column=1, padx=5, pady=5, sticky="w")

    row_index += 1  

    # Salida Estudiante y Salida Esperada 
    text_widget_estudiante = tk.Text(frame_contenedor, wrap="word", width=80)  
    text_widget_esperada = tk.Text(frame_contenedor, wrap="word", width=80)  

    text_widget_estudiante.insert("1.0", test_data[2])  # Salida Estudiante
    text_widget_esperada.insert("1.0", test_data[3])  # Salida Esperada

    # Ajustar la altura
    height_estudiante = len(test_data[2].split("\n"))
    height_esperada = len(test_data[3].split("\n"))

    text_widget_estudiante.configure(height=height_estudiante+5) # 5 lineas extras
    text_widget_esperada.configure(height=height_esperada+5)

    text_widget_estudiante.configure(state="disabled")
    text_widget_esperada.configure(state="disabled")

    text_widget_estudiante.grid(row=row_index, column=0, padx=5, pady=5, sticky="ew")
    text_widget_esperada.grid(row=row_index, column=1, padx=5, pady=5, sticky="ew")

    row_index += 1  

    # Mostrar Situacion
    tk.Label(frame_contenedor, text="Situación", font=("Arial", 12, "bold")).grid(row=row_index, column=0, padx=5, pady=5, sticky="w")
    text_widget_situacion = tk.Text(frame_contenedor, wrap="word", width=80)  # Ajustar el ancho 
    text_widget_situacion.insert("1.0", test_data[4])
    text_widget_situacion.configure(state="disabled")

    # Ajustar la altura 
    height_situacion = len(test_data[4].split("\n"))
    text_widget_situacion.configure(height=height_situacion+5)

    text_widget_situacion.grid(row=row_index, column=1, columnspan=5, padx=5, pady=5, sticky="ew")

# Etiqueta seleccionar el test
label_seleccion = tk.Label(root, text="Selecciona un test:", font=("Arial", 12))
label_seleccion.pack(pady=10)

# ComboBox seleccionar el test
combobox = ttk.Combobox(root, state="readonly", font=("Arial", 11))
combobox_values = []
i = 1
while i <= len(df_list[0]) // 5:
    combobox_values.append("TEST " + str(i))
    i += 1

combobox['values'] = combobox_values
combobox.pack(pady=10)
combobox.bind("<<ComboboxSelected>>", cargar_test)


frame_tabla = tk.Frame(root)
frame_tabla.pack(fill="both", expand=True, padx=20, pady=20)

# Canvas 
canvas = tk.Canvas(frame_tabla)
canvas.grid(row=0, column=0, sticky="nsew")  # sticky="nsew" para que ocupe todo el espacio

# scrollbar vertical al Canvas
scrollbar_vertical = ttk.Scrollbar(frame_tabla, orient="vertical", command=canvas.yview)
scrollbar_vertical.grid(row=0, column=1, sticky="ns")

# scrollbar horizontal al Canvas
scrollbar_horizontal = ttk.Scrollbar(frame_tabla, orient="horizontal", command=canvas.xview)
scrollbar_horizontal.grid(row=1, column=0, sticky="ew")

# Configurar scrollbars
canvas.configure(yscrollcommand=scrollbar_vertical.set, xscrollcommand=scrollbar_horizontal.set)

# Frame dentro del Canvas para las filas
frame_contenedor = tk.Frame(canvas)
canvas.create_window((0, 0), window=frame_contenedor, anchor="nw")

# Config scrollbar
def update_scrollregion(event):
    canvas.configure(scrollregion=canvas.bbox("all"))

# desplazamiento al evento
frame_contenedor.bind("<Configure>", update_scrollregion)

# expandir en frame_tabla
frame_tabla.grid_rowconfigure(0, weight=1)
frame_tabla.grid_columnconfigure(0, weight=1)

# evento scroll mouse
def on_mouse_wheel(event):
    canvas.yview_scroll(int(-1*(event.delta/120)), "units")  # Desplazarse con la rueda

# Asociar el evento mouse al canvas
canvas.bind_all("<MouseWheel>", on_mouse_wheel)

# empieza primer test
combobox.current(0)
cargar_test(None)

# atajenme
root.mainloop()


####################
####################
####################
################################################################################
################################################################################
################################################################################
################################################################################
################################################################################



